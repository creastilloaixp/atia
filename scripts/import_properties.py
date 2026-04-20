"""
Importador de inventario Excel → Supabase con embeddings Gemini
Fuente: A INVENTARIO Y CORRETAJE 2025.xlsx
Destino: tabla `properties` en Supabase

Uso:
  pip install openpyxl supabase requests
  python scripts/import_properties.py
"""

import os
import sys
import time
import json
import requests
import openpyxl

# ── Cargar .env.import si existe ─────────────────────────────────────────────
_env_path = os.path.join(os.path.dirname(__file__), ".env.import")
if os.path.exists(_env_path):
    with open(_env_path) as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, v = line.split("=", 1)
                os.environ.setdefault(k.strip(), v.strip())

# ── Config ───────────────────────────────────────────────────────────────────
SUPABASE_URL = os.getenv("SUPABASE_URL", "https://vjdlndntsmzoxggtruot.supabase.co")
SUPABASE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY", "")
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY", "")

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "A INVENTARIO Y CORRETAJE 2025.xlsx")
EMBED_MODEL = "models/gemini-embedding-001"
EMBED_URL = f"https://generativelanguage.googleapis.com/v1beta/{EMBED_MODEL}:embedContent?key={GEMINI_API_KEY}"

# ── Helpers ──────────────────────────────────────────────────────────────────

def get_embedding(text: str) -> list[float] | None:
    """Genera embedding de 768 dims con Gemini."""
    try:
        res = requests.post(EMBED_URL, json={
            "model": EMBED_MODEL,
            "content": {"parts": [{"text": text}]},
            "taskType": "RETRIEVAL_DOCUMENT",
        }, headers={"Content-Type": "application/json"})
        if res.status_code == 429:
            print("  Rate limit -- esperando 30s...")
            time.sleep(30)
            return get_embedding(text)  # retry
        if res.status_code != 200:
            print(f"  WARN: Embedding error {res.status_code}: {res.text[:100]}")
            return None
        return res.json()["embedding"]["values"]
    except Exception as e:
        print(f"  WARN: Embedding exception: {e}")
        return None


def supabase_insert(row: dict):
    """Inserta una propiedad en Supabase."""
    res = requests.post(
        f"{SUPABASE_URL}/rest/v1/properties",
        json=row,
        headers={
            "apikey": SUPABASE_KEY,
            "Authorization": f"Bearer {SUPABASE_KEY}",
            "Content-Type": "application/json",
            "Prefer": "return=minimal",
        },
    )
    if res.status_code not in (200, 201):
        print(f"  WARN: Insert error {res.status_code}: {res.text[:150]}")
        return False
    return True


def clean(val):
    """Limpia un valor de celda."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def parse_price(val) -> float | None:
    """Extrae precio numérico de diferentes formatos."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace(",", "").replace("$", "").replace("MXN", "").replace("mxn", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def normalize_operation(op: str | None) -> str:
    if not op:
        return "Venta"
    op = op.strip().lower()
    if "renta" in op:
        return "Renta"
    return "Venta"


def normalize_city(city: str | None) -> str:
    if not city:
        return "Culiacán"
    c = city.strip()
    if c.lower().startswith("culiac"):
        return "Culiacán"
    if c.lower().startswith("mazat"):
        return "Mazatlán"
    return c


def safe_float(val) -> float | None:
    """Convierte a float, ignorando tipos inesperados (datetime, etc)."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(str(val).strip())
    except (ValueError, TypeError):
        return None


def yes_no(val) -> bool:
    if not val:
        return False
    return str(val).strip().lower() in ("si", "sí", "sii", "yes", "s")


# ── Parsers por hoja ─────────────────────────────────────────────────────────

def parse_inv_atia(ws) -> list[dict]:
    """Parsea hoja INV. ATIA."""
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[2]:  # sin tipo de propiedad = fila vacía
            continue
        city = normalize_city(clean(row[0]))
        sector = clean(row[1])
        prop_type = clean(row[2])
        colonia = clean(row[3])
        operation = normalize_operation(clean(row[4]))
        price = parse_price(row[5])
        address = clean(row[12])
        code = clean(row[13])
        expediente = clean(row[14])

        # Texto descriptivo para RAG
        content = f"{prop_type} en {operation} en {colonia or sector or city}."
        if price:
            content += f" Precio: ${price:,.0f} MXN."
        if sector:
            content += f" Sector: {sector}."
        content += f" Ciudad: {city}."
        if address:
            content += f" Dirección: {address}."
        if expediente:
            content += f" Ref: {expediente}."

        rows.append({
            "source": "inventario_atia",
            "property_type": prop_type,
            "operation": operation,
            "city": city,
            "sector": sector,
            "colonia": colonia,
            "price": price,
            "address": address,
            "location_url": clean(row[11]),
            "images_url": clean(row[6]) if clean(row[6]) and clean(row[6]) != "Imagenes" else None,
            "lamudi_url": clean(row[7]),
            "tokko_url": clean(row[8]),
            "wiggot_url": clean(row[9]),
            "wiggot_id": clean(row[10]),
            "internal_code": code,
            "expediente": expediente,
            "content": content,
        })
    return rows


def parse_corretaje(ws) -> list[dict]:
    """Parsea hoja CORRETAJE."""
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        prop_type = clean(row[0])
        if not prop_type or prop_type.lower().startswith("culiac"):
            continue  # skip header-like rows or city-as-type
        operation = normalize_operation(clean(row[1]))
        sector = clean(row[2])
        colonia = clean(row[3])
        price = parse_price(row[4])
        bedrooms = safe_float(row[5])
        bathrooms = safe_float(row[6])
        pool = yes_no(row[7])
        private = yes_no(row[8])
        parking = safe_float(row[9])
        furnished = yes_no(row[10])
        extras = clean(row[11])
        commission = clean(row[12])
        broker = clean(row[13])
        phone = clean(row[14])
        drive = clean(row[15]) if len(row) > 15 else None

        # Texto descriptivo para RAG
        content = f"{prop_type} en {operation} en {colonia or sector or 'Culiacán'}."
        if price:
            content += f" Precio: ${price:,.0f} MXN."
        if sector:
            content += f" Sector: {sector}."
        content += " Culiacán."
        features = []
        if bedrooms and bedrooms > 0:
            features.append(f"{int(bedrooms)} recámaras")
        if bathrooms and bathrooms > 0:
            features.append(f"{bathrooms:.1f} baños")
        if parking and parking > 0:
            features.append(f"{int(parking)} estacionamientos")
        if pool:
            features.append("alberca")
        if private:
            features.append("privada")
        if furnished:
            features.append("amueblada")
        if features:
            content += " " + ", ".join(features) + "."
        if extras:
            content += f" {extras}."

        # Normalize phone
        if phone:
            phone = str(phone).replace(".0", "").strip()

        rows.append({
            "source": "corretaje",
            "property_type": prop_type.strip(),
            "operation": operation,
            "city": "Culiacán",
            "sector": sector,
            "colonia": colonia,
            "price": price,
            "bedrooms": bedrooms,
            "bathrooms": bathrooms,
            "parking_spots": parking,
            "has_pool": pool,
            "is_private": private,
            "is_furnished": furnished,
            "extra_features": extras,
            "broker_name": broker,
            "broker_phone": phone,
            "commission_pct": commission,
            "drive_link": drive,
            "content": content,
        })
    return rows


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    if not SUPABASE_KEY:
        print("ERROR: Falta SUPABASE_SERVICE_ROLE_KEY")
        print("Uso: SUPABASE_SERVICE_ROLE_KEY=xxx GEMINI_API_KEY=xxx python scripts/import_properties.py")
        sys.exit(1)
    if not GEMINI_API_KEY:
        print("ERROR: Falta GEMINI_API_KEY")
        sys.exit(1)

    print(f"Leyendo Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    # Parse both sheets
    inv_rows = parse_inv_atia(wb["INV. ATIA"])
    cor_rows = parse_corretaje(wb["CORRETAJE"])
    all_rows = inv_rows + cor_rows

    print(f"INV. ATIA: {len(inv_rows)} propiedades")
    print(f"CORRETAJE: {len(cor_rows)} propiedades")
    print(f"Total: {len(all_rows)} propiedades")
    print()

    success = 0
    errors = 0

    for i, prop in enumerate(all_rows):
        label = f"[{i+1}/{len(all_rows)}] {prop['property_type']} {prop['operation']} {prop['colonia'] or prop['sector']}"
        print(f"  {label}...")

        # Generate embedding
        emb = get_embedding(prop["content"])
        if emb:
            prop["embedding"] = emb
        else:
            print(f"    WARN: Sin embedding — se inserta sin vector")

        # Insert
        if supabase_insert(prop):
            success += 1
        else:
            errors += 1

        # Rate limiting: Gemini free tier = 1500 RPM, be safe
        if (i + 1) % 10 == 0:
            time.sleep(2)

    print()
    print(f"Importacion completa: {success} OK, {errors} errores")


if __name__ == "__main__":
    main()
